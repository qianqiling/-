import openpyxl as o

#-------------在一二次中查找---------------
names = {}#创建字典对应名字和分数
names1= []#没有重复的名单加各种信息
names2= []
def pyxl(xlsx,lim,li):    
    wb = o.load_workbook(xlsx)#接受文件名返回workbook数据类型文件
    sheet = wb.active#获取活动表 
    rows = sheet.max_row        # 获取行数
    column = sheet.max_column   # 获取列数
    #---获取姓名信息并获得没有重复姓名的名单
    for i in range(2,rows+1):                
        name = sheet.cell(i,2).value  #获取名字即第二列信息      
        names[name] = []              #利用字典做工具筛选
    #---拿出没有重复的名单---
    for i in names.keys():
        li.append(i)
    #---对名单添加默认成绩0---    
    for i in range(len(li)):
        li[i]=[li[i],0]
    #---对列表添加信息---
    for i in range(2,rows+1):
        name = sheet.cell(i,2).value  #获取名字即第二列信息
        number = sheet.cell(i,9).value#学号
        score1 = sheet.cell(i,8).value#获得分数
        time = sheet.cell(i,4).value  #做的时间
        time1=time = int(time.replace('秒',''))#对时间进行转换去掉"秒"
        if time>lim:#对超时进行判断
            score1-=(time-lim)//5#获取真实成绩
            if score1<0:
                score1=0
        for j in range(len(li)):
            if name==li[j][0]:#取多次成绩中最大值
                if score1>li[j][1]:
                    li[j][1]=score1
                    a=score1
                else:
                    a=li[j][1]
        names[name]=[name,number,a,time]
  
#--------------获得最终成绩-------------
pyxl('导论考试01--进制转换(第1次).xlsx',300,names1)
#print('第一次最终成绩为：',names1)
pyxl('导论考试01--进制转换(第2次).xlsx',300,names2)
#print('第二次最终成绩为：',names2)
#print(names)
for i in range(len(names1)):
    for j in range(len(names2)):
        if names1[i][0]==names2[j][0]:
            if names1[i][1]<names2[j][1]:
                names1[i][1]=names2[i][1]
    q=names1[i][0]
    names[q]=[names[q][0],names[q][1],names1[i][1],names[q][3]]
#print(names)    #两次的合并到第一个里
#--------------制作Excel----------------
wb1 = o.Workbook()
sh  = wb1.active
sh.title = '导论考试01--进制转换.最终成绩'
list0 = ['序号','姓名','学号','成绩']

#居中对齐
alignment = o.styles.Alignment(horizontal="center", vertical="center", text_rotation=0, wrap_text=True)
# 设置C列的宽度
sh.column_dimensions['C'].width = 12

for i in range(len(list0)):#写入标题
    sh.cell(1,i+1,value=list0[i]).alignment = alignment
for i in range(0,len(names1)):#将四列数据写入
    q=names1[i][0]
    sh.cell(i+2,1,value=i+1).alignment = alignment
    sh.cell(i+2,2,value=str(names[q][0]))
    sh.cell(i+2,3,value=str(names[q][1]))
    sh.cell(i+2,4,value=names[q][2]).alignment = alignment
wb1.save('导论考试01--进制转换.最终成绩.xlsx')    
